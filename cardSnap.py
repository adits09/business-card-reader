import os
import requests
import base64
import json
import openpyxl
from tkinter import Tk, filedialog, messagebox, Button, Label, Frame
import webbrowser
import http.server
import socketserver
import threading
import time
from pathlib import Path

API_KEY = "AIzaSyBEOC3dnhK9yNk43N46sjMP_D2r_0uFy0Y"

def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode("utf-8")

def extract_info_from_image(image_path):
    image_data = encode_image(image_path)
    url = f"https://generativelanguage.googleapis.com/v1/models/gemini-1.5-flash:generateContent?key={API_KEY}"
    headers = {
        "Content-Type": "application/json"
    }
    prompt = {
        "contents": [
            {
                "parts": [
                    {
                        "text": "Extract the Company Name, Person's Name, Designation, Phone, Email, Website, and Address from this business card. Respond only with JSON format. Use exactly these field names: 'Company Name', 'Person Name', 'Designation', 'Phone', 'Email', 'Website', 'Address'. Do NOT use triple backticks or markdown."
                    },
                    {
                        "inlineData": {
                            "mimeType": "image/jpeg",
                            "data": image_data
                        }
                    }
                ]
            }
        ]
    }
    
    response = requests.post(url, headers=headers, data=json.dumps(prompt))
    if response.status_code == 200:
        try:
            data = response.json()
            extracted_text = data['candidates'][0]['content']['parts'][0]['text']

            if "```json" in extracted_text:
                extracted_text = extracted_text.replace("```json", "").replace("```", "").strip()
            elif "```" in extracted_text:
                start_idx = extracted_text.find("```") + 3
                end_idx = extracted_text.rfind("```")
                if end_idx > start_idx:
                    extracted_text = extracted_text[start_idx:end_idx].strip()
                else:
                    extracted_text = extracted_text.replace("```", "").strip()

            print("Cleaned JSON text:", extracted_text)
            json_data = json.loads(extracted_text)

            print("Extracted data:", json_data)

            return json_data
        except json.JSONDecodeError as e:
            print("Error decoding JSON:", e)
            print("Raw response:", response.text)
            return None
        except KeyError as e:
            print(f"Error accessing key in JSON: {e}")
            print("Raw response:", response.text)
            return None
        except Exception as e:
            print("An unexpected error occurred:", e)
            print("Raw response:", response.text)
            return None
    else:
        print("API Error:", response.status_code, response.text)
        return None


def save_to_excel(info_dict, file_name="business_card_output.xlsx"):
    downloads_path = str(Path.home() / "Downloads")
    full_path = os.path.join(downloads_path, file_name)

    headers = ["Company Name", "Person Name", "Designation", "Phone", "Email", "Website", "Address"]

    field_mapping = {
        "company name": "Company Name",
        "company's name": "Company Name",
        "company": "Company Name",

        "person's name": "Person Name",
        "person name": "Person Name",
        "name": "Person Name",
        "person": "Person Name",
        "full name": "Person Name",

        "designation": "Designation",
        "title": "Designation",
        "job title": "Designation",
        "position": "Designation",
        "role": "Designation",

        "phone": "Phone",
        "phone number": "Phone",
        "mobile": "Phone",
        "telephone": "Phone",
        "tel": "Phone",
        "contact": "Phone",

        "email": "Email",
        "mail": "Email",
        "e-mail": "Email",
        "email address": "Email",

        "website": "Website",
        "web": "Website",
        "url": "Website",
        "site": "Website",
        "web address": "Website",

        "address": "Address",
        "location": "Address",
        "office": "Address",
        "office address": "Address",
    }

    print("Original data from API:", info_dict)

    normalized_data = {}
    for field, value in info_dict.items():
        normalized_field = field.lower().replace("'", "").strip()

        print(f"Processing field: '{field}' (normalized to '{normalized_field}')")

        standard_field = None
        for key, std_field in field_mapping.items():
            if normalized_field == key or key in normalized_field:
                standard_field = std_field
                print(f"  → Mapped to standard field: '{std_field}'")
                break

        if standard_field:
            normalized_data[standard_field] = value
        else:
            print(f"Warning: Unknown field '{field}' - adding as-is")
            normalized_data[field] = value

    print("Normalized data:", normalized_data)

    try:
        if os.path.isfile(full_path):
            wb = openpyxl.load_workbook(full_path)
            ws = wb.active
            for col_idx, header in enumerate(headers, start=1):
                if ws.cell(row=1, column=col_idx).value != header:
                    ws.cell(row=1, column=col_idx).value = header
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Business Cards"
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx).value = header

        next_row = ws.max_row + 1

        for col_idx, header in enumerate(headers, start=1):
            value = normalized_data.get(header, None)
            ws.cell(row=next_row, column=col_idx).value = value
            print(f"Writing '{header}': '{value}' to cell ({next_row}, {col_idx})")

        wb.save(full_path)
        print(f"Data saved to {full_path}")
        return full_path

    except Exception as e:
        print(f"Error saving file: {e}")
        try:
            fallback_path = os.path.join(os.getcwd(), file_name)
            wb.save(fallback_path)
            print(f"Data saved to fallback location: {fallback_path}")
            return fallback_path
        except Exception as e2:
            print(f"Error saving to fallback location: {e2}")
            return None


def serve_file_for_download(file_path):
    directory = os.path.dirname(file_path)
    filename = os.path.basename(file_path)

    os.chdir(directory)

    port = 8000
    handler = http.server.SimpleHTTPRequestHandler

    def run_server():
        with socketserver.TCPServer(("", port), handler) as httpd:
            print(f"Serving file at http://localhost:{port}/{filename}")
            httpd.serve_forever()

    server_thread = threading.Thread(target=run_server)
    server_thread.daemon = True
    server_thread.start()

    webbrowser.open(f"http://localhost:{port}/{filename}")

    messagebox.showinfo("Download Available",
                            f"Your file is available for download.\n\n"
                            f"It has also been saved to: {file_path}")

    time.sleep(5)
    print("Download server shutting down...")


def open_file_location(file_path):
    if os.path.exists(file_path):
        directory = os.path.dirname(file_path)
        if os.name == 'nt':
            os.startfile(directory)
        elif os.name == 'posix':
            if os.path.exists('/usr/bin/open'):
                os.system(f'open "{directory}"')
            else:
                os.system(f'xdg-open "{directory}"')

        messagebox.showinfo("File Ready",
                                f"Your Excel file is ready!\n\n"
                                f"Location: {file_path}\n\n"
                                f"The folder has been opened for you.")
    else:
        messagebox.showerror("Error", f"File not found: {file_path}")


def process_multiple_cards():
    root = Tk()
    root.title("Business Card Extractor")
    root.geometry("500x300")

    excel_path = None
    processed_count = 0
    processed_files = set()

    def select_and_process():
        nonlocal excel_path, processed_count, processed_files

        file_paths = filedialog.askopenfilenames(
            title="Select Business Card Images",
            filetypes=[("JPEG files", "*.jpg *.jpeg")]
        )

        if not file_paths:
            return

        for file_path in file_paths:
            if file_path in processed_files:
                print(f"Skipping already processed file: {os.path.basename(file_path)}")
                continue

            if not os.path.exists(file_path):
                messagebox.showerror("Error", f"File not found: {file_path}")
                continue

            status_label.config(text=f"Processing: {os.path.basename(file_path)}...")
            root.update()

            print(f"Processing: {file_path}")
            info = extract_info_from_image(file_path)

            if info:
                excel_path = save_to_excel(info)
                processed_count += 1
                processed_files.add(file_path)
                status_label.config(text=f"Processed {processed_count} card(s). Last: {os.path.basename(file_path)}")
            else:
                status_label.config(text=f"Failed to process: {os.path.basename(file_path)}")

        download_button.config(state="normal" if excel_path else "disabled")

    def download_file():
        nonlocal excel_path
        if not excel_path:
            return

        download_method = messagebox.askyesno(
            "Download Method",
            "Would you like to open the file location?\n\n"
            "Yes - Open the folder containing the Excel file\n"
            "No - Serve the file via temporary web server"
        )

        if download_method:
            open_file_location(excel_path)
        else:
            serve_file_for_download(excel_path)

    def exit_app():
        root.destroy()

    frame = Frame(root, padx=20, pady=20)
    frame.pack(expand=True, fill="both")

    Label(frame, text="Business Card Information Extractor", font=("Arial", 14, "bold")).pack(pady=10)

    Button(frame, text="Select Card Image(s)", command=select_and_process, width=20).pack(pady=10)

    status_label = Label(frame, text="Ready to process business cards...", wraplength=450)
    status_label.pack(pady=10)

    download_button = Button(frame, text="Download Excel File", command=download_file, state="disabled", width=20)
    download_button.pack(pady=10)

    Button(frame, text="Exit", command=exit_app, width=20).pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    process_multiple_cards()
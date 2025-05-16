import streamlit as st
import requests
import base64
import json
import os
import openpyxl
from pathlib import Path
import hashlib
import io
import zipfile
from PIL import Image

API_KEY = st.secrets["API_KEY"]
HEADERS = {"Content-Type": "application/json"}
MASTER_EXCEL_FILE = "business_cards_master.xlsx"

if 'processed_files' not in st.session_state:
    st.session_state.processed_files = set()

def get_file_hash(file_content):
    return hashlib.md5(file_content).hexdigest()

def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode("utf-8")

def extract_info_from_image(image_path):
    url = f"https://generativelanguage.googleapis.com/v1/models/gemini-1.5-flash:generateContent?key={API_KEY}"
    prompt = {
        "contents": [
            {
                "parts": [
                    {
                        "text": "Extract the Company Name, Person's Name, Designation, Phone, Email, Website, and Address from this business card. Respond only with JSON format. Use exactly these field names: 'Company Name', 'Person Name', 'Designation', 'Phone', 'Email', 'Website', 'Address'. Do NOT use triple backticks or markdown."
                    },
                    {
                        "inlineData": {
                            "mimeType": "image/jpeg",
                            "data": encode_image(image_path)
                        }
                    }
                ]
            }
        ]
    }
    response = requests.post(url, headers=HEADERS, data=json.dumps(prompt))
    if response.status_code == 200:
        try:
            data = response.json()
            extracted_text = data['candidates'][0]['content']['parts'][0]['text']
            for marker in ["```json", "```"]:
                extracted_text = extracted_text.replace(marker, "").strip()
            return json.loads(extracted_text)
        except Exception as e:
            st.error(f"Error processing image: {e}")
            return None
    else:
        st.error(f"API Error: {response.status_code} - {response.text}")
        return None

def normalize_fields(info_dict):
    field_mapping = {
        "company name": "Company Name",
        "company's name": "Company Name",
        "company": "Company Name",
        "person's name": "Person Name",
        "person name": "Person Name",
        "name": "Person Name",
        "person": "Person Name",
        "full name": "Person Name",
        "designation": "Designation",
        "title": "Designation",
        "job title": "Designation",
        "position": "Designation",
        "role": "Designation",
        "phone": "Phone",
        "phone number": "Phone",
        "mobile": "Phone",
        "telephone": "Phone",
        "tel": "Phone",
        "contact": "Phone",
        "email": "Email",
        "mail": "Email",
        "e-mail": "Email",
        "email address": "Email",
        "website": "Website",
        "web": "Website",
        "url": "Website",
        "site": "Website",
        "web address": "Website",
        "address": "Address",
        "location": "Address",
        "office": "Address",
        "office address": "Address",
    }
    normalized_data = {}
    for field, value in info_dict.items():
        normalized_field = field.lower().replace("'", "").strip()
        standard_field = next((std_field for key, std_field in field_mapping.items() if normalized_field == key or key in normalized_field), field)
        normalized_data[standard_field] = value
    return normalized_data

def save_to_master_excel(info_dict, file_name):
    documents_path = Path.cwd() / "documents"
    documents_path.mkdir(exist_ok=True)
    full_path = documents_path / MASTER_EXCEL_FILE
    headers = ["File Name", "Company Name", "Person Name", "Designation", "Phone", "Email", "Website", "Address"]
    normalized_data = normalize_fields(info_dict)
    
    # Add the filename as the first column
    new_row = [file_name]
    for header in headers[1:]:  # Skip "File Name" as we already added it
        new_row.append(normalized_data.get(header, ""))

    try:
        if full_path.is_file():
            wb = openpyxl.load_workbook(full_path)
            ws = wb.active
            
            # More robust duplicate checking
            existing_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_values = tuple(str(cell).strip() if cell is not None else "" for cell in row)
                existing_rows.append(row_values)
            
            new_row_tuple = tuple(str(cell).strip() if cell is not None else "" for cell in new_row)
            
            is_duplicate = False
            for existing_row in existing_rows:
                # Check email and phone match as primary keys for duplication
                # Adjust indices because we now have filename as first column
                email_match = (existing_row[5].lower() == new_row_tuple[5].lower()) and new_row_tuple[5] != ""
                phone_match = (existing_row[4].lower() == new_row_tuple[4].lower()) and new_row_tuple[4] != ""
                name_match = (existing_row[2].lower() == new_row_tuple[2].lower()) and new_row_tuple[2] != ""
                
                # Consider it duplicate if email or phone match with name
                if (email_match or phone_match) and name_match:
                    is_duplicate = True
                    break
                
                # Also check for exact row duplication (excluding the filename)
                if existing_row[1:] == new_row_tuple[1:]:
                    is_duplicate = True
                    break
            
            if not is_duplicate:
                ws.append(new_row)
                wb.save(full_path)
                return str(full_path), True
            else:
                return str(full_path), False
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Business Cards"
            ws.append(headers)
            ws.append(new_row)
            
            # Format the sheet
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
            
            wb.save(full_path)
            return str(full_path), True
    except Exception as e:
        st.error(f"Failed to save data: {e}")
        return None, False

def process_file(file_content, file_name):
    try:
        file_hash = get_file_hash(file_content)
        if file_hash in st.session_state.processed_files:
            st.info(f"Skipping already processed file: {file_name}")
            return True, False

        img = Image.open(io.BytesIO(file_content))
        temp_path = os.path.join("uploaded_cards", file_name)
        os.makedirs("uploaded_cards", exist_ok=True)
        img.save(temp_path)

        with st.spinner(f"Processing {file_name}..."):
            info = extract_info_from_image(temp_path)
            if info:
                saved_path, is_new = save_to_master_excel(info, file_name)
                if saved_path:
                    if is_new:
                        st.success(f"Data from '{file_name}' added to master Excel file!")
                    else:
                        st.info(f"Data from '{file_name}' already exists in the master Excel file.")
                    st.session_state.processed_files.add(file_hash)
                    os.remove(temp_path)
                    return True, is_new
            
            st.session_state.processed_files.add(file_hash)

        os.remove(temp_path)
        return True, False
    except Exception as e:
        st.error(f"Error processing {file_name}: {e}")
        return False, False

def main():
    st.set_page_config(page_title="CardSnap", layout="centered")
    st.title("Welcome to cardSnap")
    st.write("Upload one or more business card images (PNG, JPG, JPEG) or a ZIP file of images. We'll extract the contact info and save it to a single Excel file!")

    if st.button("Clear processed files history"):
        st.session_state.processed_files = set()
        st.success("Processing history cleared! All uploaded files will be processed again.")

    uploaded_files = st.file_uploader("Upload Business Card Images or ZIP file", type=["png", "jpg", "jpeg", "zip"], accept_multiple_files=True)

    if uploaded_files:
        new_entries_added = False
        for uploaded_file in uploaded_files:
            file_name = uploaded_file.name
            file_content = uploaded_file.read()

            if file_name.endswith(".zip"):
                try:
                    zip_hash = get_file_hash(file_content)
                    if zip_hash in st.session_state.processed_files:
                        st.info(f"Skipping already processed ZIP file: {file_name}")
                        continue

                    with zipfile.ZipFile(io.BytesIO(file_content), 'r') as zip_ref:
                        for member in zip_ref.namelist():
                            if member.lower().endswith(('.png', '.jpg', '.jpeg')):
                                with zip_ref.open(member) as image_file:
                                    image_content = image_file.read()
                                    success, is_new = process_file(image_content, os.path.basename(member))
                                    new_entries_added = new_entries_added or is_new
                    st.session_state.processed_files.add(zip_hash)
                except zipfile.BadZipFile:
                    st.error(f"Error: '{file_name}' is not a valid ZIP file.")
                except Exception as e:
                    st.error(f"Error processing '{file_name}': {e}")
            else:
                success, is_new = process_file(file_content, file_name)
                new_entries_added = new_entries_added or is_new

        # Show download button for the master Excel file
        master_excel_path = Path.cwd() / "documents" / MASTER_EXCEL_FILE
        if master_excel_path.exists():
            with open(master_excel_path, "rb") as f:
                st.download_button(
                    label=f"Download Master Contact List",
                    data=f.read(),
                    file_name=MASTER_EXCEL_FILE,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_master"
                )

        if st.session_state.processed_files:
            st.write(f"Currently processed {len(st.session_state.processed_files)} unique files.")

if __name__ == "__main__":
    main()